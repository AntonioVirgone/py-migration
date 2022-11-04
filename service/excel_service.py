from openpyxl import load_workbook

from model.card import Card


class ExcelService:
    @staticmethod
    def convertXslxToList(filename, cardVersion):
        workbook = load_workbook(filename)

        sheet = workbook.active

        cardList = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            cardList.append(Card(row[3], row[2], cardVersion))

        return cardList